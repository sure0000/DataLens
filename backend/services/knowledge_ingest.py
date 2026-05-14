"""知识条目摄取：常见办公文档 → 正文文本（Markdown 近似）。"""
from __future__ import annotations

import html
import json
import re
from html.parser import HTMLParser
from io import BytesIO
from pathlib import PurePosixPath

import httpx

MAX_INGEST_BYTES = 12 * 1024 * 1024
# 静态 HTML 抽字极少时，尝试从内嵌 SPA 状态（如 __NEXT_DATA__）补全
_MIN_BODY_BEFORE_SPA_FALLBACK = 500
_MAX_SPA_STRINGS = 450
_MAX_SPA_EXTRACT_CHARS = 800_000
# PDF 正文拼接上限（避免极大文件撑爆内存）
_MAX_PDF_EXTRACT_CHARS = 2 * 1024 * 1024


class _HtmlToTextParser(HTMLParser):
    """剥离标签，粗略保留可读文本（够用联调与普通网页）；复杂版式建议使用专用解析链。"""

    _SKIP = frozenset({"script", "style", "noscript", "template", "svg", "head"})

    def __init__(self) -> None:
        super().__init__(convert_charrefs=True)
        self._parts: list[str] = []
        self._depth_skip = 0

    def handle_starttag(self, tag: str, attrs):
        tag = tag.lower()
        if tag in self._SKIP:
            self._depth_skip += 1
            return
        if self._depth_skip:
            return
        if tag in (
            "br",
            "p",
            "div",
            "tr",
            "li",
            "h1",
            "h2",
            "h3",
            "h4",
            "h5",
            "h6",
            "article",
            "main",
            "section",
            "table",
            "thead",
            "tbody",
            "tfoot",
            "colgroup",
            "blockquote",
            "figure",
            "figcaption",
            "header",
            "footer",
            "aside",
            "hr",
            "dl",
            "dt",
            "dd",
        ):
            self._parts.append("\n")

    def handle_endtag(self, tag: str):
        tag = tag.lower()
        if tag in self._SKIP and self._depth_skip:
            self._depth_skip -= 1
            return

    def handle_data(self, data: str):
        if self._depth_skip:
            return
        chunk = html.unescape(data).strip()
        if chunk:
            self._parts.append(chunk + " ")

    def text(self) -> str:
        raw = " ".join(self._parts)
        raw = re.sub(r"[ \t]+", " ", raw)
        raw = re.sub(r"\n{3,}", "\n\n", raw)
        return raw.strip()


def _walk_json_long_strings(obj: object, acc: list[str], *, min_len: int, max_strings: int) -> None:
    if len(acc) >= max_strings:
        return
    if isinstance(obj, str):
        s = obj.strip()
        if len(s) >= min_len and any(c.isalpha() for c in s):
            acc.append(s)
        return
    if isinstance(obj, dict):
        for k, v in obj.items():
            if k in (
                "styles",
                "scriptLoader",
                "buildId",
                "assetPrefix",
                "runtimeConfig",
                "page",
                "query",
                "dynamicIds",
            ):
                continue
            _walk_json_long_strings(v, acc, min_len=min_len, max_strings=max_strings)
        return
    if isinstance(obj, list):
        for it in obj[:120]:
            _walk_json_long_strings(it, acc, min_len=min_len, max_strings=max_strings)


def _extract_text_from_spa_embedded_json(html_txt: str) -> str:
    """从 Next.js / 常见内嵌 JSON 脚本中捞取长文本，弥补纯标签解析只能得到空壳的问题。"""
    out: list[str] = []
    seen: set[str] = set()

    next_m = re.search(
        r'<script[^>]+id=["\']__NEXT_DATA__["\'][^>]*>(.*?)</script>',
        html_txt,
        re.I | re.DOTALL,
    )
    if next_m:
        raw = next_m.group(1).strip()
        try:
            data = json.loads(raw)
        except json.JSONDecodeError:
            data = None
        if isinstance(data, dict):
            buf: list[str] = []
            _walk_json_long_strings(data, buf, min_len=40, max_strings=_MAX_SPA_STRINGS)
            for s in buf:
                st = s.strip()
                if st.startswith(("http://", "https://", "//", "data:")):
                    continue
                if st in seen:
                    continue
                # 过滤明显是配置/路径串
                if st.count("/") > 12 and " " not in st:
                    continue
                seen.add(st)
                out.append(st)

    merged = "\n\n".join(out).strip()
    if len(merged) > _MAX_SPA_EXTRACT_CHARS:
        merged = merged[:_MAX_SPA_EXTRACT_CHARS] + "\n…（内嵌 JSON 提取过长已截断）"
    return merged


def _merge_html_visible_with_spa_fallback(html_txt: str, visible_plain: str) -> str:
    base = (visible_plain or "").strip()
    if len(base) >= _MIN_BODY_BEFORE_SPA_FALLBACK:
        return base
    extra = _extract_text_from_spa_embedded_json(html_txt)
    if not extra:
        return base
    note = "（以下为从页面内嵌数据尽力恢复的正文；复杂动态页仍可能不完整，可导出文件或手动粘贴）"
    if base:
        return f"{base}\n\n---\n{note}\n\n{extra}"
    return f"{note}\n\n{extra}"


def _charset_candidates_from_headers_and_html_sniff(ct_header: str, prefix: bytes) -> list[str]:
    """Collect Content-Type charset + common <meta charset=…> guesses; order-preserving."""

    cand: list[str] = []

    def _add(name: str) -> None:
        n = re.sub(r'^["\']|["\']$', "", name.strip())
        if not n:
            return
        aliases = {
            "utf8": "utf-8",
            "ascii": "utf-8",
            "cp936": "gb18030",
            "gb2312": "gb18030",
            "chinese_gbk": "gb18030",
        }
        key = aliases.get(n.lower(), n.lower())
        if key not in cand:
            cand.append(key)

    cm = re.search(r"charset=([^;\s]+)", ct_header or "", re.I)
    if cm:
        _add(cm.group(1))

    head_lc = prefix.decode("latin-1", errors="ignore").lower()
    m = re.search(r"<meta[^>]*charset\s*=\s*[\"']?\s*([^\"'\s/>]+)", head_lc)
    if m:
        _add(html.unescape(m.group(1)))
    hm = re.search(
        r"http-equiv\s*=\s*[\"']?\s*content-type[\"']?[^>]+content\s*=\s*[\"']([^\"']+)[\"']",
        head_lc,
        re.I,
    )
    if hm:
        sub = html.unescape(hm.group(1))
        sm = re.search(r"charset=([^;\s]+)", sub, re.I)
        if sm:
            _add(sm.group(1))

    _add("utf-8")
    _add("gb18030")
    return cand


def _decode_http_body_bytes(raw_bytes: bytes, ctype_header: str, sniff_html_meta: bool) -> str:
    prefix = raw_bytes[: min(len(raw_bytes), 65536)] if sniff_html_meta else b""
    cand = _charset_candidates_from_headers_and_html_sniff(ctype_header, prefix)

    seen: list[str] = []
    for c in cand:
        if c not in seen:
            seen.append(c)
    if "utf-8" not in seen:
        seen.append("utf-8")

    for enc in seen:
        try:
            return raw_bytes.decode(enc)
        except (LookupError, UnicodeDecodeError):
            continue
    return raw_bytes.decode("utf-8", errors="replace")


def _docx_body_ordered_plain(doc: object) -> str:
    """按文档中出现的顺序导出段落与表格单元格文本，弥补仅读 paragraph 遗漏表格正文的问题。"""
    from docx.oxml.ns import qn  # type: ignore[import-untyped]
    from docx.table import Table  # type: ignore[import-untyped]
    from docx.text.paragraph import Paragraph  # type: ignore[import-untyped]

    lines: list[str] = []
    body = getattr(getattr(doc, "element", None), "body", None)
    if body is None:
        return ""

    def _kids():
        ich = getattr(body, "iterchildren", None)
        if callable(ich):
            try:
                return list(ich())
            except TypeError:
                pass
        try:
            return list(body)
        except TypeError:
            return []

    for child in _kids():
        tag = getattr(child, "tag", None)
        if tag == qn("w:p"):
            tx = (Paragraph(child, doc).text or "").strip()
            if tx:
                lines.append(tx)
        elif tag == qn("w:tbl"):
            tbl = Table(child, doc)
            try:
                for row in tbl.rows:
                    cells = []
                    for cell in row.cells:
                        cx = ((cell.text or "").replace("\t", " ")).strip()
                        if cx:
                            cells.append(cx.replace("\n", " "))
                    row_txt = " | ".join(cells).strip()
                    if row_txt:
                        lines.append(row_txt)
            except Exception:
                continue

    return "\n".join(lines).strip()


ALLOWED_EXTENSIONS = frozenset({".md", ".markdown", ".txt", ".html", ".htm", ".docx", ".pdf", ".xlsx", ".csv"})


def normalize_filename(filename: str) -> str:
    name = PurePosixPath(filename.split("/")[-1]).name.strip()
    if not name:
        name = "upload.bin"
    return name


def file_to_plain(filename: str, data: bytes) -> str:
    ext = PurePosixPath(filename.lower()).suffix
    if ext not in ALLOWED_EXTENSIONS:
        raise ValueError(f"不支持的扩展名「{ext}」，允许：{', '.join(sorted(ALLOWED_EXTENSIONS))}")

    if ext in {".txt", ".md", ".markdown"}:
        return _decode_http_body_bytes(data, "", sniff_html_meta=False).strip()

    if ext in {".html", ".htm"}:
        t = _decode_http_body_bytes(data, "", sniff_html_meta=True)
        parser = _HtmlToTextParser()
        parser.feed(t)
        parser.close()
        body = _merge_html_visible_with_spa_fallback(t, parser.text()).strip()
        return body if body else t.strip()

    if ext == ".docx":
        try:
            from docx import Document  # type: ignore[import-untyped]

            doc = Document(BytesIO(data))
            merged = _docx_body_ordered_plain(doc)
            if merged:
                return merged
            return "\n".join((p.text or "").strip() for p in doc.paragraphs if (p.text or "").strip()).strip()
        except Exception as exc:
            raise ValueError(f"DOCX 解析失败：{exc}") from exc

    if ext == ".pdf":
        try:
            from pypdf import PdfReader  # type: ignore[import-untyped]

            reader = PdfReader(BytesIO(data))
            parts: list[str] = []
            total_chars = 0
            truncated = False
            for page in reader.pages:
                t = page.extract_text() or ""
                if not t.strip():
                    continue
                blk = t.strip()
                parts.append(blk)
                total_chars += len(blk)
                if total_chars >= _MAX_PDF_EXTRACT_CHARS:
                    truncated = True
                    break
            plain = "\n\n".join(parts).strip()
            if truncated:
                plain += "\n\n…（为避免超大 PDF 占用过多内存与上下文，仅在达到字符上限前的页面已导出；可按章节拆分或多条导入。）"
            return plain
        except Exception as exc:
            raise ValueError(f"PDF 解析失败：{exc}") from exc

    if ext == ".xlsx":
        try:
            from openpyxl import load_workbook  # type: ignore[import-untyped]

            wb = load_workbook(BytesIO(data), read_only=True, data_only=True)
            lines: list[str] = []
            for sn in wb.sheetnames:
                ws = wb[sn]
                lines.append(f"## {sn}")
                rows = list(ws.iter_rows(values_only=True))
                if not rows:
                    lines.append("（空工作表）")
                    continue
                # 表头
                header = [str(c or "") for c in rows[0]]
                lines.append("| " + " | ".join(header) + " |")
                lines.append("|" + "|".join(["---"] * len(header)) + "|")
                # 数据行（最多 500 行）
                for row in rows[1:501]:
                    cells = [str(c or "").replace("\n", " ").replace("|", "\\|") for c in row]
                    # 补齐到表头列数
                    while len(cells) < len(header):
                        cells.append("")
                    lines.append("| " + " | ".join(cells[:len(header)]) + " |")
                lines.append("")
            wb.close()
            return "\n".join(lines).strip()
        except Exception as exc:
            raise ValueError(f"XLSX 解析失败：{exc}") from exc

    if ext == ".csv":
        try:
            import csv
            import io

            text = _decode_http_body_bytes(data, "", sniff_html_meta=False)
            reader = csv.reader(io.StringIO(text))
            rows = list(reader)
            if not rows:
                return ""
            header = rows[0]
            lines: list[str] = ["| " + " | ".join(str(c or "") for c in header) + " |"]
            lines.append("|" + "|".join(["---"] * len(header)) + "|")
            for row in rows[1:2001]:
                cells = [str(c or "").replace("\n", " ").replace("|", "\\|") for c in row]
                while len(cells) < len(header):
                    cells.append("")
                lines.append("| " + " | ".join(cells[:len(header)]) + " |")
            return "\n".join(lines).strip()
        except Exception as exc:
            raise ValueError(f"CSV 解析失败：{exc}") from exc

    raise ValueError("未知文件类型")


def title_from_filename(filename: str) -> str:
    name = PurePosixPath(normalize_filename(filename)).stem.strip()
    return name or "未命名导入"


# ===================== 官方 API 导入（Notion 等平台） =====================

NOTION_API_VERSION = "2022-06-28"
NOTION_BASE = "https://api.notion.com/v1"


def _notion_headers(api_key: str) -> dict[str, str]:
    key = (api_key or "").strip()
    if not key:
        raise ValueError("Notion Integration Token 不能为空")
    return {
        "Authorization": f"Bearer {key}",
        "Notion-Version": NOTION_API_VERSION,
        "Content-Type": "application/json",
    }


_NOTION_ID_RE = re.compile(r"[0-9a-fA-F]{32}")


def _extract_notion_id(raw: str) -> str:
    """从 Notion URL 或原始 ID 中提取 32 位 hex 标识符（返回无横线格式）。"""
    s = (raw or "").strip()
    if "notion.so" in s.lower() or "notion.site" in s.lower():
        # 取 URL 最后一个路径段
        path = s.rstrip("/").split("/")[-1] if "/" in s else s
        # 标准格式: title-uuidwithoutdashes — 按最后一段横线拆分
        dash_parts = path.split("-")
        if len(dash_parts) > 1:
            candidate = dash_parts[-1]
            if _NOTION_ID_RE.fullmatch(candidate):
                return candidate.lower()
        # 非标准格式（无横线分隔）：从末尾反向收集 32 位 hex 字符
        hex_chars: list[str] = []
        for ch in reversed(path):
            if ch in "0123456789abcdefABCDEF":
                hex_chars.append(ch)
                if len(hex_chars) == 32:
                    return "".join(reversed(hex_chars)).lower()
        raise ValueError(f"无法从 URL 中提取 Notion 页面/数据库 ID：{raw}")
    # 纯 ID（含或不合横线）
    clean = s.replace("-", "").replace(" ", "")
    m = _NOTION_ID_RE.fullmatch(clean)
    if m:
        return m.group(0).lower()
    raise ValueError(f"无法从输入中提取 Notion 页面/数据库 ID：{raw}")


def _notion_rich_text_to_md(rt: list[dict]) -> str:
    if not rt:
        return ""
    parts: list[str] = []
    for item in rt:
        text = item.get("plain_text") or (item.get("text") or {}).get("content", "")
        ann = item.get("annotations", {})
        if ann.get("code"):
            text = f"`{text}`"
        if ann.get("bold"):
            text = f"**{text}**"
        if ann.get("italic"):
            text = f"*{text}*"
        if ann.get("strikethrough"):
            text = f"~~{text}~~"
        link = item.get("href") or (((item.get("text") or {}).get("link") or {}).get("url"))
        if link:
            text = f"[{text}]({link})"
        parts.append(text)
    return "".join(parts)


def _notion_cell_plain(cell: list[dict]) -> str:
    """将单个 table cell 的 rich_text 数组提取为纯文本。"""
    parts: list[str] = []
    for item in (cell or []):
        parts.append(item.get("plain_text") or (item.get("text") or {}).get("content", ""))
    return "".join(parts).replace("\n", " ").replace("|", "\\|")


def _render_table_rows(rows: list[dict]) -> list[str]:
    """将 table_row 列表渲染为 markdown 表格。"""
    if not rows:
        return []
    all_cells: list[list[str]] = []
    col_count = 0
    for row in rows:
        if row.get("type") != "table_row":
            continue
        cells = (row.get("table_row", {}) or {}).get("cells", [])
        row_texts = [_notion_cell_plain(c) for c in cells]
        col_count = max(col_count, len(row_texts))
        all_cells.append(row_texts)

    if not all_cells:
        return []

    lines: list[str] = []
    for ri, row_cells in enumerate(all_cells):
        # 补齐不足的列
        while len(row_cells) < col_count:
            row_cells.append("")
        lines.append("| " + " | ".join(row_cells) + " |")
        if ri == 0:
            # 表头分隔行
            lines.append("| " + " | ".join(["---"] * col_count) + " |")
    return lines


def _notion_block_to_lines(block: dict, indent: int = 0) -> list[str]:
    t = block.get("type", "")
    val = block.get(t, {}) or {}
    prefix = "  " * indent
    lines: list[str] = []

    if t in ("paragraph", "heading_1", "heading_2", "heading_3"):
        text = _notion_rich_text_to_md(val.get("rich_text", []))
        if t == "heading_1":
            lines.append(f"# {text}")
        elif t == "heading_2":
            lines.append(f"## {text}")
        elif t == "heading_3":
            lines.append(f"### {text}")
        else:
            lines.append(text or "")

    elif t == "bulleted_list_item":
        text = _notion_rich_text_to_md(val.get("rich_text", []))
        lines.append(f"{prefix}- {text}")

    elif t == "numbered_list_item":
        text = _notion_rich_text_to_md(val.get("rich_text", []))
        lines.append(f"{prefix}1. {text}")

    elif t == "to_do":
        text = _notion_rich_text_to_md(val.get("rich_text", []))
        checked = "x" if val.get("checked") else " "
        lines.append(f"{prefix}- [{checked}] {text}")

    elif t == "code":
        lang = val.get("language", "")
        text = _notion_rich_text_to_md(val.get("rich_text", []))
        lines.append(f"```{lang}\n{text}\n```")

    elif t == "quote":
        text = _notion_rich_text_to_md(val.get("rich_text", []))
        lines.append(f"{prefix}> {text}")

    elif t == "divider":
        lines.append("---")

    elif t == "child_page":
        title = val.get("title", "")
        lines.append(f"{prefix}📄 子页面：{title}")

    elif t == "table_of_contents":
        lines.append("[目录]")

    elif t == "image":
        src = (val.get("external") or val.get("file") or {}).get("url", "")
        caption = _notion_rich_text_to_md(val.get("caption", []))
        alt = caption or "图片"
        lines.append(f"{prefix}![{alt}]({src})")

    elif t == "callout":
        emoji = (b.get("callout", {}) or {}).get("icon", {}).get("emoji", "")
        text = _notion_rich_text_to_md(val.get("rich_text", []))
        prefix_icon = f"{emoji} " if emoji else ""
        lines.append(f"{prefix}> {prefix_icon}{text}")

    elif t == "bookmark":
        url = val.get("url", "")
        caption = _notion_rich_text_to_md(val.get("caption", []))
        lines.append(f"{prefix}[{caption or url}]({url})")

    elif t == "equation":
        expr = val.get("expression", "")
        lines.append(f"{prefix}${expr}$")

    elif t == "link_preview":
        url = val.get("url", "")
        lines.append(f"{prefix}> 🔗 {url}")

    elif t == "embed":
        url = val.get("url", "")
        caption = _notion_rich_text_to_md(val.get("caption", []))
        lines.append(f"{prefix}> 📎 [{caption or url}]({url})")

    elif t == "table":
        lines.append(f"{prefix}[表格]")

    elif t == "table_row":
        pass  # 由 _render_table_rows 处理

    else:
        lines.append(f"{prefix}[未支持的 Notion 块类型: {t}]")

    return lines


def fetch_official_notion_page(api_key: str, page_id: str) -> tuple[str, str]:
    """
    使用 Notion 官方 API 获取单个 Page 的完整内容（含递归子块）。
    返回 (title, markdown_text)。
    """
    headers = _notion_headers(api_key)
    clean_id = _extract_notion_id(page_id)
    # Notion API pages 端点要求带横线的 UUID
    dashed_id = f"{clean_id[:8]}-{clean_id[8:12]}-{clean_id[12:16]}-{clean_id[16:20]}-{clean_id[20:32]}"

    with httpx.Client(timeout=90.0, follow_redirects=True) as client:
        # 获取页面属性
        r = client.get(f"{NOTION_BASE}/pages/{dashed_id}", headers=headers)
        r.raise_for_status()
        page = r.json()
        if not isinstance(page, dict):
            raise ValueError(f"Notion API 返回了非预期的数据格式: {type(page).__name__}")

        # 提取标题
        title = ""
        for p in (page.get("properties") or {}).values():
            if p is not None and p.get("type") == "title":
                title = _notion_rich_text_to_md(p.get("title", []))
                break
        if not title:
            title = page.get("id", "Untitled Notion Page")

        # 递归拉取所有块（保留层级，以便处理 table / toggle 等容器块）
        def collect_blocks(block_id: str) -> list[dict]:
            result: list[dict] = []
            cursor = None
            while True:
                params: dict[str, str | int] = {"page_size": 100}
                if cursor:
                    params["start_cursor"] = cursor
                rr = client.get(
                    f"{NOTION_BASE}/blocks/{block_id}/children",
                    headers=headers,
                    params=params,
                )
                rr.raise_for_status()
                data = rr.json()
                for b in data.get("results", []):
                    kids: list[dict] = []
                    if b.get("has_children"):
                        kids = collect_blocks(b["id"])
                    b["_children"] = kids
                    result.append(b)
                if not data.get("has_more"):
                    break
                cursor = data.get("next_cursor")
            return result

        all_blocks = collect_blocks(clean_id)

        # 转文本
        def render_blocks(blocks: list[dict], indent: int = 0) -> list[str]:
            lines: list[str] = []
            for b in blocks:
                t = b.get("type", "")
                kids = b.get("_children", [])
                rendered = _notion_block_to_lines(b, indent=indent)

                if t == "table":
                    # 渲染子 table_row 为 markdown 表格
                    lines.extend(_render_table_rows(kids))
                elif t == "table_row":
                    pass  # 由父级 table 统一处理
                elif t == "toggle":
                    # 折叠块：details / summary
                    val = b.get("toggle", {}) or {}
                    summary = _notion_rich_text_to_md(val.get("rich_text", []))
                    pre = "  " * indent
                    lines.append(f"{pre}<details>")
                    lines.append(f"{pre}<summary>{summary or '(空)'}</summary>")
                    lines.append("")
                    lines.extend(render_blocks(kids, indent + 1))
                    lines.append(f"{pre}</details>")
                elif t in ("column_list", "column", "synced_block"):
                    # 透明容器：直接渲染子块
                    lines.extend(rendered)
                    lines.extend(render_blocks(kids, indent))
                else:
                    lines.extend(rendered)
                    lines.extend(render_blocks(kids, indent + 1))

            return lines

        md_lines = render_blocks(all_blocks)

        body = "\n".join(md_lines).strip()
        return title, body or "(Notion 页面内容为空或无可读取块)"


def fetch_official_notion_database(
    api_key: str, db_id: str, max_pages: int = 30
) -> list[tuple[str, str]]:
    """
    拉取 Notion Database 中的多条页面（简单实现）。
    返回 [(title, body), ...] 列表。
    """
    headers = _notion_headers(api_key)
    clean_id = _extract_notion_id(db_id)
    dashed_id = f"{clean_id[:8]}-{clean_id[8:12]}-{clean_id[12:16]}-{clean_id[16:20]}-{clean_id[20:32]}"
    results: list[tuple[str, str]] = []

    with httpx.Client(timeout=120.0) as client:
        cursor = None
        fetched = 0
        while fetched < max_pages:
            params: dict[str, str | int] = {"page_size": 10}
            if cursor:
                params["start_cursor"] = cursor
            r = client.post(
                f"{NOTION_BASE}/databases/{dashed_id}/query",
                headers=headers,
                json=params,
            )
            r.raise_for_status()
            data = r.json()
            for page in data.get("results", []):
                pid = page.get("id", "")
                if not pid:
                    continue
                try:
                    t, b = fetch_official_notion_page(api_key, pid)
                    results.append((t, b))
                    fetched += 1
                    if fetched >= max_pages:
                        break
                except Exception:
                    continue
            if not data.get("has_more") or fetched >= max_pages:
                break
            cursor = data.get("next_cursor")

    return results


# ===================== Confluence Cloud（REST API） =====================


def fetch_official_confluence_page(domain: str, email: str, api_token: str, page_id: str) -> tuple[str, str]:
    """
    Atlassian Confluence Cloud：使用邮箱 + API Token（Basic）读取页面 storage HTML。
    domain 示例：yourcompany.atlassian.net（不要含 https:// 与路径）
    page_id：页面数字 ID（非 spaceKey~title 形式）。
    """
    dom = (domain or "").strip().lower().replace("https://", "").replace("http://", "").split("/")[0].rstrip(".")
    if not dom or "." not in dom:
        raise ValueError("Confluence 请填写有效的站点域名，例如 yourcompany.atlassian.net")
    em = (email or "").strip()
    tok = (api_token or "").strip()
    if not em or not tok:
        raise ValueError("Confluence 需要邮箱与 API Token")
    pid = (page_id or "").strip()
    if not pid.isdigit():
        raise ValueError("Confluence 页面 ID 须为数字（在页面 URL 的 /pages/123456789/... 中）")

    base = f"https://{dom}/wiki/rest/api/content/{pid}"
    with httpx.Client(timeout=90.0, follow_redirects=True) as client:
        r = client.get(
            base,
            params={"expand": "body.storage,title,version"},
            auth=(em, tok),
            headers={"Accept": "application/json"},
        )
        r.raise_for_status()
        data = r.json()

    title = (data.get("title") or "").strip() or f"Confluence Page {pid}"
    storage = (((data.get("body") or {}).get("storage")) or {}).get("value") or ""
    if not storage.strip():
        return title, "（Confluence 页面无 storage 正文或无权读取 body）"

    parser = _HtmlToTextParser()
    parser.feed(storage)
    parser.close()
    body = parser.text().strip() or storage
    return title, body


# ===================== 飞书 / Lark（开放平台） =====================

FEISHU_TOKEN_URL = "https://open.feishu.cn/open-apis/auth/v3/tenant_access_token/internal"


def _feishu_tenant_access_token(app_id: str, app_secret: str) -> str:
    aid = (app_id or "").strip()
    sec = (app_secret or "").strip()
    if not aid or not sec:
        raise ValueError("飞书需要 app_id 与 app_secret（app_secret 可填在「API Key」栏）")
    with httpx.Client(timeout=45.0) as client:
        r = client.post(
            FEISHU_TOKEN_URL,
            json={"app_id": aid, "app_secret": sec},
        )
        r.raise_for_status()
        data = r.json()
    if int(data.get("code", -1)) != 0:
        raise ValueError(f"飞书获取 tenant_access_token 失败：{data.get('msg', data)}")
    token = (data.get("tenant_access_token") or "").strip()
    if not token:
        raise ValueError("飞书返回的 tenant_access_token 为空")
    return token


def _feishu_normalize_document_id(raw: str) -> str:
    s = (raw or "").strip()
    if not s:
        raise ValueError("飞书文档 ID 不能为空")
    # 从 URL 提取 docx 文档 token
    m = re.search(r"/docx/([A-Za-z0-9_-]+)", s)
    if m:
        return m.group(1)
    m2 = re.search(r"(doxcn[A-Za-z0-9_-]+)", s)
    if m2:
        return m2.group(1)
    return s


def fetch_official_feishu_doc(app_id: str, app_secret: str, document_id: str) -> tuple[str, str]:
    """
    飞书云文档：使用自建应用 app_id + app_secret 换 tenant_access_token，再拉取 docx 纯文本（Markdown 近似）。
    document_id：文档 token（doxcn…）或含 /docx/ 的完整链接。
    需在飞书开放平台为应用开通「查看、编辑和管理云空间中所有文件」等 docx 相关权限。
    """
    doc_id = _feishu_normalize_document_id(document_id)
    token = _feishu_tenant_access_token(app_id, app_secret)
    raw_url = f"https://open.feishu.cn/open-apis/docx/v1/documents/{doc_id}/raw_content"
    meta_url = f"https://open.feishu.cn/open-apis/docx/v1/documents/{doc_id}"
    with httpx.Client(timeout=90.0) as client:
        hdr = {"Authorization": f"Bearer {token}"}
        r = client.get(raw_url, headers=hdr)
        r.raise_for_status()
        data = r.json()
    if int(data.get("code", -1)) != 0:
        raise ValueError(f"飞书读取文档失败：{data.get('msg', data)}")
    inner = data.get("data") or {}
    content = (inner.get("content") or "").strip()
    title = (inner.get("title") or "").strip()
    if not title:
        with httpx.Client(timeout=45.0) as client:
            r2 = client.get(meta_url, headers={"Authorization": f"Bearer {token}"})
            r2.raise_for_status()
            d2 = r2.json()
        if int(d2.get("code", -1)) == 0:
            doc = (d2.get("data") or {}).get("document") or {}
            title = (doc.get("title") or "").strip()
    title = title or doc_id
    return title, content or "（飞书文档正文为空）"


